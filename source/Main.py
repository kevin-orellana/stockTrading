# Kevin F Orellana
# Stock Trade Profit/Loss Realization Calculator
# Notes:
#   - There are two implementations that can be used: a version using the
#       openpyxl Python library and another using the standard csv library.
#   - The openpyxl Python library can be found here:
#        To install openpyxl, I ran the command "py -m pip install openpyxl"
#        https://openpyxl.readthedocs.io/en/default/
#        This library allows high levels of float precision. However,
#        its implementation increases total runtime.
#   - The standard Python library 'csv' can also be used. However, there
#        is a significant loss of precision in realization calculations.
#        To activate the csv implementation, uncomment function
#        'csvRealizationCalculator()' on line  64

from Trader import *
import openpyxl

# calls the profit/loss realization calculator with the openpyxl implementation
def realizationCalculator():
    print("Loading xlsx file 'sampledata.xlsx' data...")
    workbook = openpyxl.load_workbook('Data/sampledata.xlsx')
    worksheet = workbook.active
    export_workbook = openpyxl.Workbook()
    export_worksheet = export_workbook.active
    dest_filename = "Data/exportdataXLSX.xlsx"
    t = Trader("badApple")
    for row in worksheet.iter_rows(min_row=1):
        # parses data from row
        date = row[0].value
        date_and_time = row[1].value
        action = row[2].value
        symbol = row[3].value
        quantity = row[4].value
        price = row[5].value
        data = [date, date_and_time, action, symbol, quantity, price]
        # updates export_list data by calling t.trade( ) function
        export_list = t.trade(data)
        # appends data to export openpyxl worksheet
        export_worksheet.append(export_list)
    # saves export_workbook to 'exportdata.xlsx'
    export_workbook.save(dest_filename)
    print("Saved processed data in ", dest_filename)

# calls the profit/loss realization calculator with the csv module implementation
def realizationCalculatorCSV():
    print("Loading csv file 'sampledata(1).csv data...")
    import_file = open("Data/sampledata(1).csv", "r")
    export_file = open("Data/exportdataCSV.csv", "w")
    t = Trader("badApple")
    for line in import_file:
        import_string = line.rstrip("\n")
        import_list = import_string.split(",")
        # updates export_list data by calling t.trade( ) function
        export_list = t.trade(import_list)
        export_string = ",".join(export_list) + "\n"
        # appends data to export_file csv file
        export_file.write(export_string)

    import_file.close()
    export_file.close()
    print("Data exported to exportdata(1).csv")
if __name__ == "__main__":
    realizationCalculator()
    # realizationCalculatorCSV()
